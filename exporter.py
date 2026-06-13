import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


def _style_header(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _write_table(ws, start_row: int, headers: list, rows: list) -> int:
    for i, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=i, value=h)
    _style_header(ws, start_row, len(headers))
    for r_idx, row in enumerate(rows, start_row + 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN_BORDER
    return start_row + len(rows) + 2


def _auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def build_report(
    domain_key: str,
    domain_label: str,
    start: str,
    end: str,
    gsc_summary: dict,
    gsc_queries: list,
    gsc_pages: list,
    ga4_summary: dict,
    ga4_pages: list,
    ga4_sources: list,
    meta_campaigns: list = None,
) -> bytes:
    wb = Workbook()

    # --- GSC Sheet ---
    ws = wb.active
    ws.title = "Search Console"
    ws.cell(row=1, column=1, value=f"{domain_label} — Search Console Report").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Period: {start} to {end}").font = Font(italic=True, color="666666")

    row = 4
    headers = ["Metric", "Value"]
    data = [
        ("Total Clicks", gsc_summary.get("clicks", 0)),
        ("Total Impressions", gsc_summary.get("impressions", 0)),
        ("Avg CTR (%)", gsc_summary.get("ctr", 0)),
        ("Avg Position", gsc_summary.get("position", 0)),
    ]
    row = _write_table(ws, row, headers, data)

    row = _write_table(ws, row, ["Query", "Clicks", "Impressions", "CTR (%)", "Position"],
                        [(q["query"], q["clicks"], q["impressions"], q["ctr"], q["position"]) for q in gsc_queries])

    row = _write_table(ws, row, ["Page", "Clicks", "Impressions", "CTR (%)", "Position"],
                        [(p["page"], p["clicks"], p["impressions"], p["ctr"], p["position"]) for p in gsc_pages])
    _auto_width(ws)

    # --- GA4 Sheet ---
    if ga4_summary:
        ws2 = wb.create_sheet("Analytics")
        ws2.cell(row=1, column=1, value=f"{domain_label} — GA4 Analytics Report").font = Font(bold=True, size=14)
        ws2.cell(row=2, column=1, value=f"Period: {start} to {end}").font = Font(italic=True, color="666666")

        row = 4
        data = [
            ("Sessions", ga4_summary.get("sessions", 0)),
            ("Total Users", ga4_summary.get("users", 0)),
            ("New Users", ga4_summary.get("new_users", 0)),
            ("Bounce Rate (%)", ga4_summary.get("bounce_rate", 0)),
            ("Avg Session Duration (s)", ga4_summary.get("avg_session", 0)),
            ("Pageviews", ga4_summary.get("pageviews", 0)),
        ]
        row = _write_table(ws2, row, ["Metric", "Value"], data)

        row = _write_table(ws2, row, ["Page", "Views", "Sessions", "Avg Duration (s)"],
                            [(p["page"], p["views"], p["sessions"], p["avg_dur"]) for p in ga4_pages])

        row = _write_table(ws2, row, ["Channel", "Sessions", "Users"],
                            [(s["channel"], s["sessions"], s["users"]) for s in ga4_sources])
        _auto_width(ws2)

    # --- Meta Sheet ---
    if meta_campaigns:
        ws3 = wb.create_sheet("Meta Ads")
        ws3.cell(row=1, column=1, value=f"{domain_label} — Meta Ads Report").font = Font(bold=True, size=14)
        ws3.cell(row=2, column=1, value=f"Period: {start} to {end}").font = Font(italic=True, color="666666")

        row = 4
        row = _write_table(ws3, row,
                            ["Campaign", "Status", "Spend", "Impressions", "Reach", "Clicks", "CTR (%)", "CPC"],
                            [(c["name"], c["status"], c["spend"], c["impressions"], c["reach"],
                              c["clicks"], c["ctr"], c["cpc"]) for c in meta_campaigns])
        _auto_width(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
