import csv
import io
from openpyxl import load_workbook


def _try_number(val):
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return val
    s = str(val).strip().replace(",", "").replace("%", "")
    try:
        return float(s) if "." in s else int(s)
    except (ValueError, TypeError):
        return val


def _parse_xlsx(file_bytes: bytes):
    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h or f"col_{i}").strip() for i, h in enumerate(next(rows_iter, []))]
    rows = []
    for row in rows_iter:
        if all(c is None for c in row):
            continue
        rows.append({headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))})
    wb.close()
    return headers, rows


def _parse_csv(file_bytes: bytes):
    text = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    headers = reader.fieldnames or []
    rows = list(reader)
    return list(headers), rows


def _summarize(headers, rows):
    summary = {"total_rows": len(rows)}
    for h in headers:
        vals = [_try_number(r.get(h)) for r in rows]
        nums = [v for v in vals if isinstance(v, (int, float))]
        if nums and len(nums) > len(rows) * 0.5:
            summary[h] = {"sum": round(sum(nums), 2), "avg": round(sum(nums) / len(nums), 2)}
    return summary


def parse_linkedin_analytics(file_bytes: bytes, filename: str = "") -> dict:
    if filename.lower().endswith(".csv"):
        headers, rows = _parse_csv(file_bytes)
    else:
        headers, rows = _parse_xlsx(file_bytes)

    for row in rows:
        for k in row:
            row[k] = _try_number(row[k])

    return {
        "headers": headers,
        "rows": rows[:100],
        "summary": _summarize(headers, rows),
    }


def parse_linkedin_followers(file_bytes: bytes, filename: str = "") -> dict:
    return parse_linkedin_analytics(file_bytes, filename)
